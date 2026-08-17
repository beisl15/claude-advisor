#!/usr/bin/env python3
"""
Serie trimestral da ASML direto do RI oficial (asml.com), nao de agregador.
Atualiza a entrada ASML em data/fin.json. Requer: openpyxl.

POR QUE ESTE SCRIPT EXISTE
--------------------------
A ASML e foreign private issuer: so arquiva 20-F anual na SEC, sem 10-Q, entao
nao ha serie trimestral em XBRL. O fetch_intl.py usava o Yahoo, que devolve no
maximo 5 trimestres -- teto do servidor, confirmado em 11/08/2026 pedindo janelas
de 2, 6 e 12 anos (sempre 5). Aqui a fonte e a propria companhia.

COMO A ASML PUBLICA (verificado em 11/08/2026)
---------------------------------------------
Cada trimestre tem uma pagina de URL previsivel:

    https://www.asml.com/en/investors/financial-results/q{N}-{ANO}

A pagina e server-rendered (o link aparece no HTML sem executar JavaScript) e
contem um XLSX oficial cujo endereco tem um GUID nao adivinhavel:

    https://ourbrand.asml.com/asset/<guid>/Financial-statements-US-GAAP-Q{N}-{ANO}-excel.xlsx

Por isso o fluxo e: baixa a pagina -> acha o link por regex -> baixa a planilha.

O detalhe que torna isso barato: cada arquivo e um "Quarterly Summary" com CINCO
trimestres (colunas C, E, G, I, K), nao so o trimestre do titulo. Quatro arquivos
espacados de um ano cobrem os 17 trimestres com 4 downloads em vez de 17.

Layout das abas (constante entre arquivos):
  linha 9   -> datas de fim de trimestre em serial Excel, nas colunas C/E/G/I/K
  linha 12+ -> rotulo na coluna A, valores nas mesmas colunas
  abas      -> 'Q Statements of Operations', 'Q Ratios and Other Data',
               'Q Consolidated Balance Sheets', 'Q Statements of Cash Flows'

Valores em milhoes de euros -> saem em bilhoes, como o resto do dashboard.
"""
import datetime
import io
import json
import os
import re
import sys
import urllib.request

try:
    from openpyxl import load_workbook
except ImportError:
    print("!! openpyxl nao instalado (pip install openpyxl)", file=sys.stderr)
    raise

HERE = os.path.dirname(__file__)
OUT = os.path.join(HERE, "..", "data", "fin.json")
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124 Safari/537.36")

PAGINA = "https://www.asml.com/en/investors/financial-results/q{q}-{y}"
NQ = 17
COLS = ("C", "E", "G", "I", "K")   # colunas de trimestre dentro da planilha
LINHA_DATAS = 9

# rotulo na planilha -> chave no nosso FIN. Casamento e por prefixo normalizado,
# porque a ASML as vezes deixa espaco sobrando no fim do rotulo.
OPS = {
    "total net sales": "rev",
    "gross profit": "gp",
    "income from operations": "ebitda",   # EBIT (ver nota do dashboard)
    "net income": "ni",
    "diluted net income per ordinary share": "eps",
}
BAL = {
    "cash and cash equivalents": "_cash",
    "short-term investments": "_inv",
    "current portion of long-term debt": "_debt_c",
    "long-term debt": "_debt_nc",
}
CF = {
    "purchase of property, plant and equipment": "capex",
    "dividend paid": "div",
    "purchase of treasury shares": "buyback",
}


def baixa(url, binario=False):
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=90) as r:
        return r.read() if binario else r.read().decode("utf-8", "replace")


def link_planilha(q, y):
    """Acha o XLSX na pagina do trimestre. GUID nao e adivinhavel -> tem de raspar."""
    html = baixa(PAGINA.format(q=q, y=y))
    pat = (r"https://ourbrand\.asml\.com/asset/[a-f0-9-]+/"
           r"Financial-statements-US-GAAP-Q%d-%d-excel\.xlsx" % (q, y))
    m = re.search(pat, html, re.I)
    return m.group(0) if m else None


def serial_para_data(v):
    """Serial do Excel -> 'YYYY-MM-DD'. Epoca do Excel e 1899-12-30.

    Atencao: a ASML usa calendario de 52/53 semanas, entao os trimestres fecham
    em DOMINGO e nao no ultimo dia do mes -- Q1 2026 termina em 2026-03-29, nao
    em 31/03. So o exercicio fecha sempre em 31/12. Nao "arredondar" essas datas
    para fim de mes: o rotulo do trimestre sai do mes (ver qlabel), e mexer nisso
    desalinharia a serie com o resto do dashboard.
    """
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    try:
        n = int(float(v))
    except (TypeError, ValueError):
        return None
    if not (20000 < n < 60000):     # fora disso nao e data plausivel
        return None
    return (datetime.date(1899, 12, 30) + datetime.timedelta(days=n)).strftime("%Y-%m-%d")


def norm(s):
    return re.sub(r"\s+", " ", str(s or "")).strip().lower().rstrip(":")


def le_aba(ws, mapa):
    """Devolve {chave: {data: valor}} para os rotulos pedidos nesta aba."""
    datas = {}
    for c in COLS:
        d = serial_para_data(ws[f"{c}{LINHA_DATAS}"].value)
        if d:
            datas[c] = d
    if not datas:
        return {}
    out = {}
    for row in ws.iter_rows(min_row=LINHA_DATAS + 1):
        rotulo = norm(row[0].value)
        if not rotulo:
            continue
        chave = mapa.get(rotulo)
        if chave is None:
            continue
        for c, d in datas.items():
            v = ws[f"{c}{row[0].row}"].value
            if isinstance(v, (int, float)):
                out.setdefault(chave, {})[d] = float(v)
    return out


def le_planilha(url):
    wb = load_workbook(io.BytesIO(baixa(url, binario=True)), data_only=True)
    dados = {}
    for nome in wb.sheetnames:
        n = nome.lower()
        if "operations" in n:
            mapa = OPS
        elif "balance" in n:
            mapa = BAL
        elif "cash flow" in n:
            mapa = CF
        else:
            continue
        for k, serie in le_aba(wb[nome], mapa).items():
            dados.setdefault(k, {}).update(serie)
    return dados


def qlabel(d):
    y, m = int(d[:4]), int(d[5:7])
    return f"CY{y}Q{(m - 1) // 3 + 1}", f"Q{(m - 1) // 3 + 1}'{d[2:4]}"


def ancoras():
    """Trimestres-ancora: o atual e um por ano anterior. Cada arquivo traz 5
    trimestres, entao 4-5 ancoras cobrem os 17 com folga e poucos downloads."""
    hoje = datetime.date.today()
    q, y = (hoje.month - 1) // 3 + 1, hoje.year
    # recua um trimestre: o mais recente pode ainda nao ter sido publicado
    q -= 1
    if q == 0:
        q, y = 4, y - 1
    return [(q, y - k) for k in range(0, 5)]


def main():
    dados = {}
    for q, y in ancoras():
        try:
            url = link_planilha(q, y)
        except Exception as ex:
            print(f". pagina Q{q} {y}: {ex}", file=sys.stderr)
            continue
        if not url:
            print(f". Q{q} {y}: XLSX nao encontrado na pagina", file=sys.stderr)
            continue
        try:
            for k, serie in le_planilha(url).items():
                dados.setdefault(k, {}).update(serie)
            print(f"ok Q{q} {y}")
        except Exception as ex:
            print(f". planilha Q{q} {y}: {ex}", file=sys.stderr)

    if not dados.get("rev"):
        print("!! ASML: nenhuma receita extraida — nada gravado", file=sys.stderr)
        return 1

    # eixo de datas = trimestres com receita, um por trimestre civil
    bykey = {}
    for d in sorted(dados["rev"]):
        bykey[qlabel(d)[0]] = d
    datas = [bykey[k] for k in sorted(bykey)][-NQ:]

    def col(k, escala=1e3):     # milhoes -> bilhoes
        s = dados.get(k, {})
        return [round(s[d] / escala, 3) if d in s else None for d in datas]

    def bruto(k):
        return dados.get(k, {})

    netdebt = []
    for d in datas:
        div = bruto("_debt_c").get(d), bruto("_debt_nc").get(d)
        if div[0] is None and div[1] is None:
            netdebt.append(None)
            continue
        cx = (bruto("_cash").get(d) or 0) + (bruto("_inv").get(d) or 0)
        netdebt.append(round(((div[0] or 0) + (div[1] or 0) - cx) / 1e3, 3))

    entry = dict(
        src="ASML IR (US GAAP quarterly summary)", asof=qlabel(datas[-1])[1], moeda="EUR",
        q=[qlabel(d)[1] for d in datas],
        rev=col("rev"), gp=col("gp"), ebitda=col("ebitda"), ni=col("ni"),
        eps=col("eps", 1),               # ja vem por acao
        capex=[abs(v) if v is not None else None for v in col("capex")],
        div=[abs(v) if v is not None else None for v in col("div")],
        buyback=[abs(v) if v is not None else None for v in col("buyback")],
        netdebt=netdebt,
    )

    try:
        with open(OUT) as f:
            fin = json.load(f)
    except Exception:
        fin = {}
    fin["ASML"] = entry
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w") as f:
        json.dump(fin, f, indent=1, ensure_ascii=False)
    print(f"-> ASML: {entry['q'][0]}..{entry['q'][-1]} ({len(datas)} trimestres) "
          f"rev_ult={entry['rev'][-1]}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
